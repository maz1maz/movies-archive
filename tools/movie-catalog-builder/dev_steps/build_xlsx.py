import csv, json, re, sys
sys.path.insert(0, '/home/user/work')
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

W = '/home/user/work/'
rows = list(csv.DictReader(open(W + 'fixed.csv', encoding='utf-8')))
picked = json.load(open(W + 'picked_final.json'))
meta = json.load(open(W + 'meta.json'))
wd = json.load(open(W + 'wd.json'))
syn = json.load(open(W + 'syn.json'))
names = meta['names']
ratings = meta['ratings']
prin = meta['prin']

COUNTRY_FIX = {
    'United States of America': 'USA', 'United States': 'USA',
    'United Kingdom': 'UK', 'Kingdom of the Netherlands': 'Netherlands',
    "People's Republic of China": 'China', 'Republic of Korea': 'South Korea',
    'Soviet Union': 'USSR', 'Federal Republic of Germany': 'West Germany',
    'German Democratic Republic': 'East Germany', 'Republic of Ireland': 'Ireland',
    'Kingdom of Denmark': 'Denmark', 'Czech Republic': 'Czechia',
    'Hong Kong': 'Hong Kong', 'Republic of China': 'Taiwan',
}

MPA_FIX = {
    'R (MPAA film rating)': 'R', 'PG-13 (MPAA film rating)': 'PG-13',
    'PG (MPAA film rating)': 'PG', 'G (MPAA film rating)': 'G',
    'NC-17 (MPAA film rating)': 'NC-17', 'X (MPAA film rating)': 'X',
    'M (MPAA film rating)': 'M', 'GP (MPAA film rating)': 'GP',
    'Not Rated': 'NR', 'Unrated': 'NR', 'Approved': 'Approved',
    'MPAA film rating system: R': 'R',
}
MPA_OK = {'G', 'PG', 'PG-13', 'R', 'NC-17', 'X', 'M', 'GP', 'NR', 'Approved', 'Passed', 'TV-MA', 'TV-14', 'TV-PG', 'TV-G'}


def clean_mpa(lst):
    out = []
    for m in lst or []:
        m = MPA_FIX.get(m, m)
        m = re.sub(r'\s*\(.*?\)\s*', '', m).strip()
        m = re.sub(r'^MPAA?\s*(film\s*)?rating\s*:?\s*', '', m, flags=re.I).strip()
        if m in MPA_OK:
            out.append(m)
    # prefer canonical modern order
    for pref in ['G', 'PG', 'PG-13', 'R', 'NC-17', 'X', 'M', 'GP', 'Approved', 'Passed', 'NR']:
        if pref in out:
            return pref
    return out[0] if out else ''


def clean_country(lst):
    out = []
    for c in lst or []:
        c = COUNTRY_FIX.get(c, c)
        if c not in out:
            out.append(c)
    return ', '.join(out[:4])


def cast_of(t, n=5):
    lst = prin.get(t) or []
    lst = sorted(lst, key=lambda x: x[0])
    seen, out = set(), []
    for _, nm in lst:
        who = names.get(nm)
        if who and who not in seen:
            seen.add(who)
            out.append(who)
        if len(out) >= n:
            break
    return ', '.join(out)


def directors_of(t):
    return ', '.join(names[n] for n in meta['crew'].get(t, []) if n in names)


def synopsis(t):
    s = syn.get(t) or {}
    e = (s.get('extract') or '').strip()
    e = re.sub(r'\s+', ' ', e)
    return e


def poster(t):
    s = syn.get(t) or {}
    w = wd.get(t) or {}
    return w.get('image') or s.get('thumb') or ''


def orig_title(t, current):
    w = wd.get(t) or {}
    c = picked_rec(t)
    cands = []
    if c:
        if c['orig'] and c['orig'] != c['prim']:
            cands.append(c['orig'])
    for o in (w.get('orig') or []):
        cands.append(o)
    cur = re.sub(r'\W+', '', (current or '')).lower()
    for x in cands:
        x = (x or '').strip()
        if len(x) < 3:
            continue
        if re.sub(r'\W+', '', x).lower() == cur:
            continue
        return x
    return ''


_recmap = {}
for k, v in picked.items():
    if v and v.get('c'):
        _recmap[v['c']['t']] = v['c']


def picked_rec(t):
    return _recmap.get(t)


HEADERS = ["Title", "Original Title", "Shelf", "Row", "Director", "Cast", "Year", "Genre",
           "Rating (IMDb)", "Runtime (min)", "Country", "Synopsis", "Poster URL", "Studio",
           "MPA Rating", "Format", "Watched", "My Rating (1-5)", "Criterion (yes/no)", "Copies",
           "Media Type (Physical/Digital)", "Drive Number", "Content Type (Movie/Series)",
           "Seasons/Episodes", "IMDb ID", "IMDb Link"]

KEEP_EMPTY = {"Shelf", "Row", "Watched", "My Rating (1-5)", "Criterion (yes/no)", "Copies",
              "Media Type (Physical/Digital)", "Drive Number", "Content Type (Movie/Series)",
              "Seasons/Episodes"}

wb = Workbook()
ws = wb.active
ws.title = "Collection"
ws.append(HEADERS)

stats = {h: 0 for h in HEADERS}
unmatched = []

for r in rows:
    i = r['#']
    p = picked.get(i)
    c = p['c'] if p and p.get('c') else None
    t = c['t'] if c else None
    w = wd.get(t) or {} if t else {}

    title = r['Title'].strip()
    out = {h: '' for h in HEADERS}
    out['Title'] = title
    # preserve user's own values
    out['Shelf'] = r.get('Shelf', '')
    out['Content Type (Movie/Series)'] = ''
    out['Genre'] = r.get('Genre', '')
    out['Year'] = r.get('Year', '')
    out['Director'] = r.get('Director', '')

    if c:
        out['IMDb ID'] = t
        out['IMDb Link'] = f'https://www.imdb.com/title/{t}/'
        out['Original Title'] = orig_title(t, title)
        d = directors_of(t)
        if d:
            out['Director'] = d
        out['Cast'] = cast_of(t)
        if c['y'] and c['y'] != r'\N':
            out['Year'] = c['y']
        g = (c['g'] or '').replace(r'\N', '')
        if g:
            out['Genre'] = ', '.join(g.split(','))
        if t in ratings:
            out['Rating (IMDb)'] = float(ratings[t][0])
        rt = c['rt']
        if rt and rt != r'\N':
            out['Runtime (min)'] = int(rt)
        out['Country'] = clean_country(w.get('country')) or r.get('Country', '')
        out['Synopsis'] = synopsis(t)
        out['Poster URL'] = poster(t)
        out['Studio'] = ', '.join((w.get('studio') or [])[:3])
        out['MPA Rating'] = clean_mpa(w.get('mpa'))
    else:
        unmatched.append((i, title))
        out['Rating (IMDb)'] = float(r['Rating (IMDb)']) if r.get('Rating (IMDb)') else ''
        out['Runtime (min)'] = int(r['Runtime (min)']) if r.get('Runtime (min)') else ''
        out['Country'] = r.get('Country', '')

    for h in KEEP_EMPTY - {'Shelf'}:
        out[h] = ''
    ws.append([out[h] for h in HEADERS])
    for h in HEADERS:
        if out[h] != '':
            stats[h] += 1

# ---- styling ----
hdr_fill = PatternFill("solid", fgColor="8EA9DB")
thin = Side(style="thin", color="D9E1F2")
WIDTHS = {"Title": 34, "Original Title": 28, "Shelf": 7, "Row": 6, "Director": 24, "Cast": 42,
          "Year": 7, "Genre": 24, "Rating (IMDb)": 9, "Runtime (min)": 9, "Country": 20,
          "Synopsis": 80, "Poster URL": 40, "Studio": 28, "MPA Rating": 9, "Format": 9,
          "Watched": 9, "My Rating (1-5)": 9, "Criterion (yes/no)": 11, "Copies": 8,
          "Media Type (Physical/Digital)": 15, "Drive Number": 11,
          "Content Type (Movie/Series)": 15, "Seasons/Episodes": 14,
          "IMDb ID": 12, "IMDb Link": 34}

for i, h in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=i)
    cell.font = Font(bold=True, size=10, color="1F2937")
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = WIDTHS[h]
ws.row_dimensions[1].height = 34
ws.freeze_panes = "C2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"

idx = {h: i + 1 for i, h in enumerate(HEADERS)}
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.border = Border(bottom=thin)
        cell.alignment = Alignment(vertical="top", wrap_text=False)
    row[idx['Synopsis'] - 1].alignment = Alignment(vertical="top", wrap_text=True)
    row[idx['Cast'] - 1].alignment = Alignment(vertical="top", wrap_text=True)
    for h in ('Year', 'Rating (IMDb)', 'Runtime (min)', 'MPA Rating', 'Shelf', 'Row', 'Copies'):
        row[idx[h] - 1].alignment = Alignment(horizontal="center", vertical="top")
    row[idx['Rating (IMDb)'] - 1].number_format = '0.0'
    lk = row[idx['IMDb Link'] - 1]
    if lk.value:
        lk.hyperlink = lk.value
        lk.font = Font(color="0563C1", underline="single", size=10)
    pu = row[idx['Poster URL'] - 1]
    if pu.value:
        pu.hyperlink = pu.value
        pu.font = Font(color="0563C1", underline="single", size=10)

# dropdowns for the user-managed columns
def dv(col, formula):
    d = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
    ws.add_data_validation(d)
    L = get_column_letter(idx[col])
    d.add(f"{L}2:{L}{ws.max_row}")

dv('Watched', '"Yes,No"')
dv('Criterion (yes/no)', '"Yes,No"')
dv('Media Type (Physical/Digital)', '"Physical,Digital,Both"')
dv('Content Type (Movie/Series)', '"Movie,Series"')
dv('My Rating (1-5)', '"1,2,3,4,5"')
dv('Format', '"Blu-ray,4K UHD,DVD,VHS,Digital,Other"')

wb.save('/home/user/Movie_Catalog_Filled.xlsx')

print("rows:", ws.max_row - 1)
print("unmatched:", len(unmatched))
for h in HEADERS:
    print(f"  {h:32} {stats[h]}")
json.dump(unmatched, open(W + 'unmatched.json', 'w'))

# ---------- extra sheets ----------
from openpyxl.styles import Font as F2
ws2 = wb.create_sheet("Needs Review")
ws2.append(["Row #", "Title", "Reason"])
BOXNOTE = {
 '190':'Box set (Alain Robbe-Grillet: Six Films) - not a single title',
 '367':'Double feature (Arthur + Arthur 2)',
 '759':'Criterion box set (Bo Widerberg)',
 '925':'Box set (Charlie Chaplin Collection)',
 '935':'Box set (Avant-Garde anthology)',
 '121':'Abbott & Costello collection - multiple films',
}
for i, t in unmatched:
    ws2.append([i, t, BOXNOTE.get(i, 'No confident IMDb match - please verify')])
for c in ws2[1]:
    c.font = F2(bold=True); c.fill = hdr_fill
ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 55

ws3 = wb.create_sheet("Summary")
ws3.append(["Column", "Filled", "Empty", "% Filled"])
total = ws.max_row - 1
for h in HEADERS:
    f = stats[h]
    ws3.append([h, f, total - f, round(100.0 * f / total, 1)])
for c in ws3[1]:
    c.font = F2(bold=True); c.fill = hdr_fill
ws3.column_dimensions['A'].width = 32
for col in 'BCD':
    ws3.column_dimensions[col].width = 11
ws3.append([])
ws3.append(["Total rows", total])
ws3.append(["Matched to IMDb", stats['IMDb ID']])
ws3.append(["Source", "IMDb official datasets + Wikidata + Wikipedia"])
ws3.append(["Built", "2026-07-25"])

wb.save('/home/user/Movie_Catalog_Filled.xlsx')
print("saved with extra sheets")

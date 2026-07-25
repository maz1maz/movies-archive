#!/usr/bin/env python3
"""
Movie Catalog Builder
=====================
Fills a movie-collection spreadsheet with metadata pulled from free, no-API-key
sources:

  * IMDb official datasets (https://datasets.imdbws.com) -> title, year, genre,
    runtime, rating, director, cast
  * Wikidata (via the QLever mirror)                     -> original title,
    country, studio, poster, MPA rating, Wikipedia link
  * Wikipedia REST API                                   -> synopsis, poster fallback

These columns are intentionally LEFT EMPTY for the user to fill by hand:
  Shelf*, Row, Watched, My Rating, Criterion, Copies, Media Type,
  Drive Number, Content Type, Seasons/Episodes
  (*Shelf is preserved from the input file if present.)

Usage
-----
    python3 catalog_builder.py --input movies.csv --output catalog.xlsx

    # resume-friendly: every stage caches to --workdir and is skipped if done
    python3 catalog_builder.py -i movies.csv -o out.xlsx --workdir ./cache

    # re-run only the spreadsheet build after tweaking styling
    python3 catalog_builder.py -i movies.csv -o out.xlsx --only build

Input CSV must have at least a `Title` column. `Year` and `Director` are
optional but massively improve match accuracy. A `#` id column is used if
present, otherwise row numbers are assigned.

Requires: openpyxl   (pip install openpyxl)
"""

import argparse
import csv
import gzip
import json
import os
import re
import sys
import time
import unicodedata
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor

# --------------------------------------------------------------------------
# Config
# --------------------------------------------------------------------------

IMDB_FILES = ["title.basics", "title.ratings", "title.crew",
              "title.principals", "name.basics", "title.akas"]
IMDB_URL = "https://datasets.imdbws.com/{}.tsv.gz"

WD_ENDPOINT = "https://qlever.cs.uni-freiburg.de/api/wikidata/"
WIKI_SUMMARY = "https://en.wikipedia.org/api/rest_v1/page/summary/"
UA = "MovieCatalogBuilder/1.0 (personal film collection catalog)"

# title types worth considering as a "film in a collection"
GOOD_TYPES = ("movie", "tvMovie", "video", "tvSpecial", "short",
              "tvSeries", "tvMiniSeries")
TYPE_WEIGHT = {"movie": 4, "tvMovie": 2, "tvMiniSeries": 2, "tvSeries": 1,
               "video": 0, "tvSpecial": 0, "short": -3}

HEADERS = ["Title", "Original Title", "Shelf", "Row", "Director", "Cast", "Year",
           "Genre", "Rating (IMDb)", "Runtime (min)", "Country", "Synopsis",
           "Poster URL", "Studio", "MPA Rating", "Format", "Watched",
           "My Rating (1-5)", "Criterion (yes/no)", "Copies",
           "Media Type (Physical/Digital)", "Drive Number",
           "Content Type (Movie/Series)", "Seasons/Episodes",
           "IMDb ID", "IMDb Link"]

# columns the user fills in manually -> always written blank
USER_COLUMNS = {"Row", "Watched", "My Rating (1-5)", "Criterion (yes/no)",
                "Copies", "Media Type (Physical/Digital)", "Drive Number",
                "Content Type (Movie/Series)", "Seasons/Episodes"}

# Titles in the source list that are typos / alternate names.
# Map: row id -> better search title. Edit for your own collection.
TITLE_FIXES = {
    '46': 'Un coeur en hiver', '135': 'Acasa, My Home',
    '164': 'Adela jeste nevecerela', '186': 'The Wild Pear Tree',
    '222': 'All Summers End', '224': 'All the Colors of the Dark',
    '310': 'Anita', '639': 'Bill & Ted Face the Music',
    '669': 'Black Dahlia', '670': 'The Black Dahlia',
    '727': 'Blood and Black Lace', '818': 'The Boys Next Door',
    '881': 'The Brothers Grimsby', '924': 'Cars',
    '933': 'All About Lily Chou-Chou',
    '934': 'Marina Abramovic: The Artist Is Present',
    '139': 'Accident', '367': 'Arthur', '148': 'Adoption',
    '121': 'Buck Privates', '61': 'Pope Francis: A Man of His Word',
    '578': 'My Journey Through French Cinema', '459': 'Land of Mine',
    '932': "Bluebeard's Eighth Wife", '304': 'Angst',
    '722': 'Blood and Flowers', '671': 'The Sun at Midnight',
}

# Rows that are box sets / compilations -> deliberately not matched.
BOX_SETS = {
    '190': 'Box set (Alain Robbe-Grillet: Six Films) - not a single title',
    '367': 'Double feature (Arthur + Arthur 2)',
    '759': 'Criterion box set (Bo Widerberg)',
    '925': 'Box set (Charlie Chaplin Collection)',
    '935': 'Box set (Avant-Garde anthology)',
    '121': 'Abbott & Costello collection - multiple films',
}

# Hand-verified IMDb ids where automatic matching picks the wrong film.
ID_OVERRIDES = {
    '182': 'tt1780967',    # Seberg (released as "Against All Enemies")
    '207': 'tt0103644',    # Alien 3  (dataset spells it Alien³)
    '224': 'tt0069390',    # Tutti i colori del buio
    '648': 'tt7713068',    # Birds of Prey (Harley Quinn)
    '779': 'tt0080464',    # The Boogey Man (1980, Ulli Lommel)
    '828': 'tt10410506',   # Brainwashed: Sex-Camera-Power
    '157': 'tt7711170',    # Alone (2020, John Hyams)
}

COUNTRY_FIX = {
    'United States of America': 'USA', 'United States': 'USA',
    'United Kingdom': 'UK', 'Kingdom of the Netherlands': 'Netherlands',
    "People's Republic of China": 'China', 'Republic of Korea': 'South Korea',
    'Soviet Union': 'USSR', 'Federal Republic of Germany': 'West Germany',
    'German Democratic Republic': 'East Germany',
    'Republic of Ireland': 'Ireland', 'Kingdom of Denmark': 'Denmark',
    'Czech Republic': 'Czechia', 'Republic of China': 'Taiwan',
}

MPA_VALID = {'G', 'PG', 'PG-13', 'R', 'NC-17', 'X', 'M', 'GP', 'NR',
             'Approved', 'Passed', 'TV-MA', 'TV-14', 'TV-PG', 'TV-G'}
MPA_ORDER = ['G', 'PG', 'PG-13', 'R', 'NC-17', 'X', 'M', 'GP',
             'Approved', 'Passed', 'NR']


# --------------------------------------------------------------------------
# Text helpers
# --------------------------------------------------------------------------

def strip_accents(s):
    s = unicodedata.normalize('NFKD', s)
    return ''.join(c for c in s if not unicodedata.combining(c))


def norm(s):
    """Aggressive title key: lowercase, no accents/punctuation, no leading article."""
    s = strip_accents(s).lower()
    for a, b in (('&', ' and '), ('ß', 'ss'), ('ø', 'o'), ('æ', 'ae'),
                 ('œ', 'oe'), ('ð', 'd'), ('þ', 'th'), ('ł', 'l')):
        s = s.replace(a, b)
    s = re.sub(r"[^a-z0-9?]+", " ", s).strip()
    s = re.sub(r"^(the|a|an|le|la|les|el|il|un|une|der|die|das|los|las) ", "", s)
    return re.sub(r"\s+", " ", s)


def variants(title):
    """All plausible lookup keys for one messy catalog title."""
    t = title.strip()
    base = [t]
    m = re.match(r"^(.*),\s*(The|A|An|Der|Die|Das|Le|La|Les|Il|El)$", t, re.I)
    if m:                                    # "Andromeda Strain, The"
        base.append("%s %s" % (m.group(2), m.group(1)))

    out = set()
    for b in base:
        cand = [b]
        c = re.sub(r"\s*\((?:copy|disc|vol\.?|volume|part)\s*\d+\)\s*$", "", b, flags=re.I)
        c = re.sub(r"\s*\(\d{4}\)\s*$", "", c)
        cand.append(c)
        if '/' in c:                          # "A Scandal in Paris / Lured"
            cand += [p.strip() for p in c.split('/')]
        m2 = re.match(r"^(.*?)\s*\((.+)\)\s*$", c)
        if m2:                                # "Afire (Roter Himmel)"
            cand += [m2.group(1).strip(), m2.group(2).strip()]
        m3 = re.match(r"^(.*?)\s+([12])$", c)
        if m3:                                # "Avengers 1"
            cand.append(m3.group(1).strip())
        for x in cand:
            n = norm(x)
            if len(n) > 1:
                out.add(n)
    return out


def name_key(s):
    return re.sub(r"[^a-z ]+", " ", strip_accents(s).lower()).strip()


def director_set(s):
    out = set()
    for x in re.split(r'[,/&]', (s or '').replace(' and ', ',')):
        x = name_key(x)
        if x and x not in ('various directors', 'criterion', 'various'):
            out.add(x)
    return out


def same_person(a, b):
    if a == b:
        return True
    pa, pb = a.split(), b.split()
    if not pa or not pb:
        return False
    if pa[-1] == pb[-1] and pa[0][:1] == pb[0][:1]:
        return True
    return pa[-1] == pb[-1] and (len(pa) == 1 or len(pb) == 1)


def read_rows(path):
    """Read the catalog CSV, auto-repairing legacy DOS/Windows encodings."""
    raw = open(path, 'rb').read()
    text = None
    for enc in ('utf-8-sig', 'utf-8', 'cp1252', 'cp437', 'latin-1'):
        try:
            cand = raw.decode(enc)
        except UnicodeDecodeError:
            continue
        # cp437 mojibake check: real accents beat replacement junk
        if enc in ('utf-8-sig', 'utf-8'):
            text = cand
            break
        text = cand
        break
    if text is None:
        text = raw.decode('latin-1')
    # If the utf-8 decode failed earlier the file is legacy: prefer cp437,
    # which is what DOS-era exports use and renders accents correctly.
    try:
        raw.decode('utf-8')
    except UnicodeDecodeError:
        text = raw.decode('cp437')

    rows = list(csv.DictReader(text.splitlines()))
    for n, r in enumerate(rows, 1):
        if not r.get('#'):
            r['#'] = str(n)
    return rows


# --------------------------------------------------------------------------
# Stage 1 - download IMDb datasets
# --------------------------------------------------------------------------

def stage_download(wd):
    d = os.path.join(wd, 'imdb')
    os.makedirs(d, exist_ok=True)
    for f in IMDB_FILES:
        dst = os.path.join(d, f + '.tsv.gz')
        if os.path.exists(dst) and os.path.getsize(dst) > 1000:
            print("  have %s" % f)
            continue
        print("  downloading %s ..." % f, flush=True)
        req = urllib.request.Request(IMDB_URL.format(f), headers={'User-Agent': UA})
        with urllib.request.urlopen(req, timeout=600) as r, open(dst, 'wb') as o:
            while True:
                chunk = r.read(1 << 20)
                if not chunk:
                    break
                o.write(chunk)
    return d


def tsv(path):
    with gzip.open(path, 'rt', encoding='utf-8', errors='replace', newline='') as f:
        rd = csv.reader(f, delimiter='\t', quoting=csv.QUOTE_NONE)
        next(rd, None)
        for row in rd:
            yield row


# --------------------------------------------------------------------------
# Stage 2 - candidate titles
# --------------------------------------------------------------------------

def stage_candidates(rows, imdb, wd):
    out = os.path.join(wd, 'candidates.json')
    if os.path.exists(out):
        print("  cached")
        return json.load(open(out))

    want = {}
    for r in rows:
        titles = {r['Title']}
        if r['#'] in TITLE_FIXES:
            titles.add(TITLE_FIXES[r['#']])
        for t in titles:
            for v in variants(t):
                want.setdefault(v, set()).add(r['#'])
    print("  %d lookup keys" % len(want))

    cands = {}
    for row in tsv(os.path.join(imdb, 'title.basics.tsv.gz')):
        if len(row) < 9 or row[1] not in GOOD_TYPES:
            continue
        for t in {row[2], row[3]}:
            n = norm(t)
            if n in want:
                cands.setdefault(n, []).append(
                    {'t': row[0], 'type': row[1], 'prim': row[2], 'orig': row[3],
                     'y': row[5], 'rt': row[7], 'g': row[8]})
                break
    print("  %d keys matched in title.basics" % len(cands))

    # AKA pass for whatever is still missing (foreign release titles)
    missing = set(want) - set(cands)
    if missing:
        print("  AKA scan for %d unresolved keys (slow) ..." % len(missing), flush=True)
        hits = {}
        for row in tsv(os.path.join(imdb, 'title.akas.tsv.gz')):
            if len(row) < 4:
                continue
            n = norm(row[2])
            if n in missing:
                hits.setdefault(n, set()).add(row[0])
        need = {t for v in hits.values() for t in v}
        recs = {}
        for row in tsv(os.path.join(imdb, 'title.basics.tsv.gz')):
            if len(row) >= 9 and row[0] in need and row[1] in GOOD_TYPES:
                recs[row[0]] = {'t': row[0], 'type': row[1], 'prim': row[2],
                                'orig': row[3], 'y': row[5], 'rt': row[7], 'g': row[8]}
        added = 0
        for k, ts in hits.items():
            for t in ts:
                if t in recs:
                    cands.setdefault(k, []).append(recs[t])
                    added += 1
        print("  +%d candidates from AKAs" % added)

    json.dump(cands, open(out, 'w'))
    return cands


# --------------------------------------------------------------------------
# Stage 3 - crew / cast / ratings for candidates
# --------------------------------------------------------------------------

def stage_meta(cands, imdb, wd):
    out = os.path.join(wd, 'meta.json')
    if os.path.exists(out):
        print("  cached")
        return json.load(open(out))

    ids = {c['t'] for v in cands.values() for c in v} | set(ID_OVERRIDES.values())
    print("  %d candidate titles" % len(ids))

    ratings, crew, prin, need = {}, {}, {}, set()
    for r in tsv(os.path.join(imdb, 'title.ratings.tsv.gz')):
        if r[0] in ids:
            ratings[r[0]] = (r[1], int(r[2]))
    for r in tsv(os.path.join(imdb, 'title.crew.tsv.gz')):
        if r[0] in ids:
            d = [x for x in r[1].split(',') if x.startswith('nm')]
            crew[r[0]] = d
            need.update(d)
    for r in tsv(os.path.join(imdb, 'title.principals.tsv.gz')):
        if r[0] in ids and r[3] in ('actor', 'actress', 'self'):
            prin.setdefault(r[0], []).append((int(r[1]), r[2]))
            need.add(r[2])
    names = {}
    for r in tsv(os.path.join(imdb, 'name.basics.tsv.gz')):
        if r[0] in need:
            names[r[0]] = r[1]

    meta = {'ratings': ratings, 'crew': crew, 'prin': prin, 'names': names}
    print("  ratings=%d crew=%d cast=%d names=%d"
          % (len(ratings), len(crew), len(prin), len(names)))
    json.dump(meta, open(out, 'w'))
    return meta


# --------------------------------------------------------------------------
# Stage 4 - pick the right IMDb title per row
# --------------------------------------------------------------------------

def score(cand, want_year, want_dirs, exact_keys, meta):
    s = TYPE_WEIGHT.get(cand['type'], 0)
    if norm(cand['prim']) in exact_keys or norm(cand['orig']) in exact_keys:
        s += 3
    try:
        y = int(cand['y'])
    except (ValueError, TypeError):
        y = None
    if want_year and y:
        d = abs(want_year - y)
        s += 9 if d == 0 else (5 if d == 1 else (2 if d <= 2 else -7))
    elif want_year and not y:
        s -= 2
    dirs = {name_key(meta['names'][n]) for n in meta['crew'].get(cand['t'], [])
            if n in meta['names']}
    if want_dirs and dirs:
        s += 12 if any(same_person(a, b) for a in want_dirs for b in dirs) else -6
    elif want_dirs and not dirs:
        s -= 1
    r = meta['ratings'].get(cand['t'])
    s += (2 + min(4, r[1] ** 0.5 / 120)) if r else -1
    return s


def stage_match(rows, cands, meta, imdb, wd):
    out = os.path.join(wd, 'picked.json')
    if os.path.exists(out):
        print("  cached")
        return json.load(open(out))

    # fetch basics for hand-set overrides
    ovr_recs = {}
    if ID_OVERRIDES:
        need = set(ID_OVERRIDES.values())
        for row in tsv(os.path.join(imdb, 'title.basics.tsv.gz')):
            if row[0] in need:
                ovr_recs[row[0]] = {'t': row[0], 'type': row[1], 'prim': row[2],
                                    'orig': row[3], 'y': row[5], 'rt': row[7],
                                    'g': row[8]}

    picked, weak = {}, []
    for r in rows:
        rid = r['#']
        if rid in BOX_SETS:
            picked[rid] = None
            continue
        if rid in ID_OVERRIDES and ID_OVERRIDES[rid] in ovr_recs:
            picked[rid] = {'c': ovr_recs[ID_OVERRIDES[rid]], 'score': 99}
            continue

        titles = {r['Title']}
        if rid in TITLE_FIXES:
            titles.add(TITLE_FIXES[rid])
        pool = {}
        for t in titles:
            for v in variants(t):
                for c in cands.get(v, []):
                    pool[c['t']] = c
        if not pool:
            picked[rid] = None
            weak.append((rid, r['Title'], 'no candidates'))
            continue

        try:
            wy = int(r.get('Year') or 0) or None
        except ValueError:
            wy = None
        wd_dirs = director_set(r.get('Director'))
        keys = {norm(t) for t in titles}
        best, bs = None, -1e9
        for c in pool.values():
            s = score(c, wy, wd_dirs, keys, meta)
            if s > bs:
                bs, best = s, c
        picked[rid] = {'c': best, 'score': round(bs, 2)}
        if bs < 8:
            weak.append((rid, r['Title'], 'low confidence: %s' % best['prim']))

    ok = sum(1 for v in picked.values() if v)
    print("  matched %d/%d   (%d weak, %d box sets)"
          % (ok, len(rows), len(weak), len(BOX_SETS)))
    json.dump(picked, open(out, 'w'))
    json.dump(weak, open(os.path.join(wd, 'weak.json'), 'w'))
    return picked


# --------------------------------------------------------------------------
# Stage 5 - Wikidata enrichment
# --------------------------------------------------------------------------

WD_QUERY = '''PREFIX wdt: <http://www.wikidata.org/prop/direct/>
PREFIX rdfs: <http://www.w3.org/2000/01/rdf-schema#>
PREFIX schema: <http://schema.org/>
SELECT ?imdb ?orig ?cl ?sl ?image ?ml ?enwiki WHERE {
  VALUES ?imdb { %s }
  ?item wdt:P345 ?imdb .
  OPTIONAL { ?item wdt:P1476 ?orig }
  OPTIONAL { ?item wdt:P495 ?c . ?c rdfs:label ?cl . FILTER(lang(?cl)="en") }
  OPTIONAL { ?item wdt:P272 ?s . ?s rdfs:label ?sl . FILTER(lang(?sl)="en") }
  OPTIONAL { ?item wdt:P3383 ?image }
  OPTIONAL { ?item wdt:P1657 ?m . ?m rdfs:label ?ml . FILTER(lang(?ml)="en") }
  OPTIONAL { ?enwiki schema:about ?item . ?enwiki schema:isPartOf <https://en.wikipedia.org/> }
}'''


def stage_wikidata(picked, wd, workers=4, batch=25):
    out = os.path.join(wd, 'wikidata.json')
    done = json.load(open(out)) if os.path.exists(out) else {}
    ids = sorted({v['c']['t'] for v in picked.values() if v and v.get('c')})
    todo = [i for i in ids if i not in done]
    print("  %d to fetch (%d cached)" % (len(todo), len(done)))
    if not todo:
        return done

    def fetch(chunk):
        vals = ' '.join('"%s"' % c for c in chunk)
        url = WD_ENDPOINT + '?' + urllib.parse.urlencode({'query': WD_QUERY % vals})
        for a in range(3):
            try:
                req = urllib.request.Request(
                    url, headers={'User-Agent': UA,
                                  'Accept': 'application/sparql-results+json'})
                return chunk, json.loads(urllib.request.urlopen(req, timeout=180).read())
            except Exception:
                time.sleep(2 * (a + 1))
        return chunk, None

    def g(b, k):
        return b[k]['value'] if k in b else None

    chunks = [todo[i:i + batch] for i in range(0, len(todo), batch)]
    n = 0
    with ThreadPoolExecutor(max_workers=workers) as ex:
        for chunk, res in ex.map(fetch, chunks):
            n += len(chunk)
            if res is None:
                continue
            agg = {}
            for b in res['results']['bindings']:
                i = g(b, 'imdb')
                if not i:
                    continue
                a = agg.setdefault(i, {'orig': set(), 'country': set(),
                                       'studio': set(), 'image': None,
                                       'mpa': set(), 'wiki': None})
                if g(b, 'orig'):
                    a['orig'].add(g(b, 'orig'))
                if g(b, 'cl'):
                    a['country'].add(g(b, 'cl'))
                if g(b, 'sl'):
                    a['studio'].add(g(b, 'sl'))
                if g(b, 'image') and not a['image']:
                    a['image'] = g(b, 'image')
                if g(b, 'ml'):
                    a['mpa'].add(g(b, 'ml'))
                if g(b, 'enwiki') and not a['wiki']:
                    a['wiki'] = g(b, 'enwiki')
            for i in chunk:
                a = agg.get(i)
                done[i] = None if not a else {
                    'orig': sorted(a['orig']), 'country': sorted(a['country']),
                    'studio': sorted(a['studio']), 'image': a['image'],
                    'mpa': sorted(a['mpa']), 'wiki': a['wiki']}
            json.dump(done, open(out, 'w'))
            print("    %d/%d" % (n, len(todo)), flush=True)
    return done


# --------------------------------------------------------------------------
# Stage 6 - Wikipedia synopses (rate-limited; run repeatedly)
# --------------------------------------------------------------------------

def stage_wikipedia(picked, wikidata, wd, workers=2, passes=4):
    out = os.path.join(wd, 'synopsis.json')
    done = json.load(open(out)) if os.path.exists(out) else {}
    done = {k: v for k, v in done.items() if v}          # drop earlier failures
    ids = sorted({v['c']['t'] for v in picked.values() if v and v.get('c')})

    def fetch(task):
        tid, page = task
        url = WIKI_SUMMARY + urllib.parse.quote(page, safe='')
        for a in range(3):
            try:
                req = urllib.request.Request(url, headers={'User-Agent': UA})
                d = json.loads(urllib.request.urlopen(req, timeout=45).read())
                img = (d.get('originalimage') or d.get('thumbnail') or {}).get('source')
                return tid, {'extract': d.get('extract'), 'thumb': img}
            except Exception as e:
                if getattr(e, 'code', None) == 404:
                    return tid, None
                time.sleep(1.5 * (a + 1))
        return tid, None

    for p in range(passes):
        tasks = []
        for i in ids:
            if i in done:
                continue
            w = (wikidata.get(i) or {}).get('wiki')
            if w:
                tasks.append((i, urllib.parse.unquote(w.rsplit('/', 1)[-1])))
        if not tasks:
            break
        print("  pass %d: %d pages" % (p + 1, len(tasks)), flush=True)
        n = 0
        with ThreadPoolExecutor(max_workers=workers) as ex:
            for tid, res in ex.map(fetch, tasks):
                n += 1
                if res:
                    done[tid] = res
                if n % 50 == 0:
                    json.dump(done, open(out, 'w'))
                    print("    %d/%d" % (n, len(tasks)), flush=True)
        json.dump(done, open(out, 'w'))
        print("  have %d synopses" % len(done))
    return done


# --------------------------------------------------------------------------
# Stage 7 - build the spreadsheet
# --------------------------------------------------------------------------

def clean_mpa(values):
    got = []
    for m in values or []:
        m = re.sub(r'\s*\(.*?\)\s*', '', m).strip()
        m = re.sub(r'^MPAA?\s*(film\s*)?rating\s*:?\s*', '', m, flags=re.I).strip()
        if m in ('Not Rated', 'Unrated'):
            m = 'NR'
        if m in MPA_VALID:
            got.append(m)
    for pref in MPA_ORDER:
        if pref in got:
            return pref
    return got[0] if got else ''


def clean_country(values):
    out = []
    for c in values or []:
        c = COUNTRY_FIX.get(c, c)
        if c not in out:
            out.append(c)
    return ', '.join(out[:4])


def build_workbook(rows, picked, meta, wikidata, synopsis, outfile):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    names, crew, prin, ratings = (meta['names'], meta['crew'],
                                  meta['prin'], meta['ratings'])

    def cast_of(t, n=5):
        seen, out = set(), []
        for _, nm in sorted(prin.get(t) or [], key=lambda x: x[0]):
            who = names.get(nm)
            if who and who not in seen:
                seen.add(who)
                out.append(who)
            if len(out) >= n:
                break
        return ', '.join(out)

    def original_title(rec, wdrec, current):
        cands = []
        if rec and rec['orig'] and rec['orig'] != rec['prim']:
            cands.append(rec['orig'])
        cands += (wdrec.get('orig') or [])
        cur = re.sub(r'\W+', '', current or '').lower()
        for x in cands:
            x = (x or '').strip()
            if len(x) >= 3 and re.sub(r'\W+', '', x).lower() != cur:
                return x
        return ''

    wb = Workbook()
    ws = wb.active
    ws.title = "Collection"
    ws.append(HEADERS)

    stats = dict.fromkeys(HEADERS, 0)
    unmatched = []

    for r in rows:
        rid = r['#']
        p = picked.get(rid)
        rec = p['c'] if p and p.get('c') else None
        t = rec['t'] if rec else None
        w = (wikidata.get(t) or {}) if t else {}

        o = dict.fromkeys(HEADERS, '')
        o['Title'] = r['Title'].strip()
        o['Shelf'] = r.get('Shelf', '') or ''
        o['Year'] = r.get('Year', '')
        o['Genre'] = r.get('Genre', '')
        o['Director'] = r.get('Director', '')

        if rec:
            o['IMDb ID'] = t
            o['IMDb Link'] = 'https://www.imdb.com/title/%s/' % t
            o['Original Title'] = original_title(rec, w, o['Title'])
            d = ', '.join(names[n] for n in crew.get(t, []) if n in names)
            if d:
                o['Director'] = d
            o['Cast'] = cast_of(t)
            if rec['y'] and rec['y'] != r'\N':
                o['Year'] = rec['y']
            g = (rec['g'] or '').replace(r'\N', '')
            if g:
                o['Genre'] = ', '.join(g.split(','))
            if t in ratings:
                o['Rating (IMDb)'] = float(ratings[t][0])
            if rec['rt'] and rec['rt'] != r'\N':
                o['Runtime (min)'] = int(rec['rt'])
            o['Country'] = clean_country(w.get('country')) or r.get('Country', '')
            s = synopsis.get(t) or {}
            o['Synopsis'] = re.sub(r'\s+', ' ', (s.get('extract') or '').strip())
            o['Poster URL'] = w.get('image') or s.get('thumb') or ''
            o['Studio'] = ', '.join((w.get('studio') or [])[:3])
            o['MPA Rating'] = clean_mpa(w.get('mpa'))
        else:
            unmatched.append((rid, o['Title'],
                              BOX_SETS.get(rid, 'no confident IMDb match')))
            o['Rating (IMDb)'] = float(r['Rating (IMDb)']) if r.get('Rating (IMDb)') else ''
            o['Runtime (min)'] = int(r['Runtime (min)']) if r.get('Runtime (min)') else ''
            o['Country'] = r.get('Country', '')

        for h in USER_COLUMNS:
            o[h] = ''
        ws.append([o[h] for h in HEADERS])
        for h in HEADERS:
            if o[h] != '':
                stats[h] += 1

    # ---- styling ----
    widths = {"Title": 34, "Original Title": 28, "Shelf": 7, "Row": 6,
              "Director": 24, "Cast": 42, "Year": 7, "Genre": 24,
              "Rating (IMDb)": 9, "Runtime (min)": 9, "Country": 20,
              "Synopsis": 80, "Poster URL": 40, "Studio": 28, "MPA Rating": 9,
              "Format": 9, "Watched": 9, "My Rating (1-5)": 9,
              "Criterion (yes/no)": 11, "Copies": 8,
              "Media Type (Physical/Digital)": 15, "Drive Number": 11,
              "Content Type (Movie/Series)": 15, "Seasons/Episodes": 14,
              "IMDb ID": 12, "IMDb Link": 34}
    hdr_fill = PatternFill("solid", fgColor="8EA9DB")
    thin = Side(style="thin", color="D9E1F2")

    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, size=10, color="1F2937")
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(i)].width = widths[h]
    ws.row_dimensions[1].height = 34
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(HEADERS)), ws.max_row)

    idx = {h: i + 1 for i, h in enumerate(HEADERS)}
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for c in row:
            c.border = Border(bottom=thin)
            c.alignment = Alignment(vertical="top")
        for h in ('Synopsis', 'Cast'):
            row[idx[h] - 1].alignment = Alignment(vertical="top", wrap_text=True)
        for h in ('Year', 'Rating (IMDb)', 'Runtime (min)', 'MPA Rating',
                  'Shelf', 'Row', 'Copies'):
            row[idx[h] - 1].alignment = Alignment(horizontal="center", vertical="top")
        row[idx['Rating (IMDb)'] - 1].number_format = '0.0'
        for h in ('IMDb Link', 'Poster URL'):
            c = row[idx[h] - 1]
            if c.value:
                c.hyperlink = c.value
                c.font = Font(color="0563C1", underline="single", size=10)

    def dropdown(col, formula):
        dv = DataValidation(type="list", formula1=formula, allow_blank=True,
                            showDropDown=False)
        ws.add_data_validation(dv)
        L = get_column_letter(idx[col])
        dv.add("%s2:%s%d" % (L, L, ws.max_row))

    dropdown('Watched', '"Yes,No"')
    dropdown('Criterion (yes/no)', '"Yes,No"')
    dropdown('Media Type (Physical/Digital)', '"Physical,Digital,Both"')
    dropdown('Content Type (Movie/Series)', '"Movie,Series"')
    dropdown('My Rating (1-5)', '"1,2,3,4,5"')
    dropdown('Format', '"Blu-ray,4K UHD,DVD,VHS,Digital,Other"')

    # ---- Needs Review ----
    ws2 = wb.create_sheet("Needs Review")
    ws2.append(["Row #", "Title", "Reason"])
    for u in unmatched:
        ws2.append(list(u))
    for c in ws2[1]:
        c.font = Font(bold=True)
        c.fill = hdr_fill
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['C'].width = 55

    # ---- Summary ----
    ws3 = wb.create_sheet("Summary")
    ws3.append(["Column", "Filled", "Empty", "% Filled"])
    total = ws.max_row - 1
    for h in HEADERS:
        ws3.append([h, stats[h], total - stats[h],
                    round(100.0 * stats[h] / total, 1) if total else 0])
    for c in ws3[1]:
        c.font = Font(bold=True)
        c.fill = hdr_fill
    ws3.column_dimensions['A'].width = 32
    for col in 'BCD':
        ws3.column_dimensions[col].width = 11
    ws3.append([])
    ws3.append(["Total rows", total])
    ws3.append(["Matched to IMDb", stats['IMDb ID']])
    ws3.append(["Sources", "IMDb datasets + Wikidata + Wikipedia"])
    ws3.append(["Built", time.strftime('%Y-%m-%d')])

    wb.save(outfile)
    return stats, unmatched, total


# --------------------------------------------------------------------------
# Main
# --------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser(description="Fill a movie catalog spreadsheet.")
    ap.add_argument('-i', '--input', required=True, help='source CSV')
    ap.add_argument('-o', '--output', default='catalog.xlsx', help='output .xlsx')
    ap.add_argument('-w', '--workdir', default='./cache', help='cache directory')
    ap.add_argument('--only', choices=['download', 'match', 'web', 'build'],
                    help='run a single stage (uses cached data for the rest)')
    ap.add_argument('--wiki-passes', type=int, default=4,
                    help='Wikipedia retry rounds (rate limits cause partial fetches)')
    a = ap.parse_args()

    os.makedirs(a.workdir, exist_ok=True)
    rows = read_rows(a.input)
    print("Loaded %d rows from %s" % (len(rows), a.input))

    print("\n[1/6] IMDb datasets")
    imdb = stage_download(a.workdir)
    if a.only == 'download':
        return

    print("\n[2/6] Candidate titles")
    cands = stage_candidates(rows, imdb, a.workdir)

    print("\n[3/6] Crew, cast and ratings")
    meta = stage_meta(cands, imdb, a.workdir)

    print("\n[4/6] Matching")
    picked = stage_match(rows, cands, meta, imdb, a.workdir)
    if a.only == 'match':
        return

    print("\n[5/6] Wikidata")
    wikidata = stage_wikidata(picked, a.workdir)

    print("\n[6/6] Wikipedia synopses")
    synopsis = stage_wikipedia(picked, wikidata, a.workdir, passes=a.wiki_passes)
    if a.only == 'web':
        return

    print("\nBuilding workbook ...")
    stats, unmatched, total = build_workbook(rows, picked, meta, wikidata,
                                             synopsis, a.output)
    print("\nSaved %s" % a.output)
    print("  %d rows, %d matched to IMDb, %d need review"
          % (total, stats['IMDb ID'], len(unmatched)))
    for h in HEADERS:
        if h not in USER_COLUMNS and h != 'Shelf':
            print("    %-30s %4d  (%.0f%%)"
                  % (h, stats[h], 100.0 * stats[h] / total if total else 0))


if __name__ == '__main__':
    main()
